"""
스마트스토어 경쟁사 가격 모니터
PyInstaller exe 변환 가능 단일 스크립트
"""
import sys, os, json, time, threading, re, queue
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from pathlib import Path
from datetime import datetime

# 필수 패키지 자동 설치
def ensure_packages():
    pkgs = {"requests": "requests", "bs4": "beautifulsoup4",
            "openpyxl": "openpyxl"}
    import importlib, subprocess
    for mod, pkg in pkgs.items():
        try:
            importlib.import_module(mod)
        except ImportError:
            subprocess.check_call([sys.executable, "-m", "pip",
                                   "install", pkg, "-q"])

ensure_packages()

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# 경로 설정 (exe 변환 대응)
def base_dir():
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).parent

BASE       = base_dir()
CONFIG_F   = BASE / "config.json"
HISTORY_F  = BASE / "history.json"
LICENSE_F  = BASE / "license.json"
REPORTS_D  = BASE / "reports"
REPORTS_D.mkdir(exist_ok=True)

PLANS = {
    "BASIC":    {"label": "베이직",   "limit": 1, "permanent": False, "months": 3, "price": 59000},
    "STANDARD": {"label": "스탠다드", "limit": 3, "permanent": False, "months": 3, "price": 139000},
    "PREMIUM":  {"label": "프리미엄", "limit": 5, "permanent": True,  "months": 0, "price": 359000},
}

def verify_key(raw_key: str) -> dict | None:
    """
    Public GitHub build: proprietary license verification is intentionally removed.
    Use the private build for real customer license validation.
    """
    return None

def demo_license() -> dict:
    plan = "STANDARD"
    return {
        "plan": plan,
        "limit": PLANS[plan]["limit"],
        "label": f"{PLANS[plan]['label']} 데모",
        "expires": "공개 데모",
        "permanent": False,
        "valid": True,
    }

def load_license() -> dict | None:
    if os.environ.get("SSTORE_PUBLIC_DEMO", "1") == "1":
        return demo_license()
    if not LICENSE_F.exists():
        return None
    try:
        data = json.loads(LICENSE_F.read_text(encoding="utf-8"))
        info = verify_key(data.get("key", ""))
        if info is None or not info.get("valid"):
            try:
                LICENSE_F.unlink()
            except:
                pass
            return None
        return info
    except:
        return None

def save_license(key: str) -> bool:
    return False
# 색상/폰트
C = {
    "bg":     "#0f0f0f", "panel":  "#181818",
    "card":   "#222222", "border": "#2d2d2d",
    "green":  "#03C75A", "red":    "#FF4747",
    "yellow": "#FFA500", "blue":   "#4A9EFF",
    "text":   "#f0f0f0", "sub":    "#777",
    "input":  "#1a1a1a",
}
FONT = ("Malgun Gothic", 10)
FONT_B = ("Malgun Gothic", 10, "bold")
FONT_T = ("Malgun Gothic", 16, "bold")
FONT_M = ("Consolas", 9)
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
    "Accept-Language": "ko-KR,ko;q=0.9,en;q=0.8",
}

# 네이버 쇼핑 검색
# 1순위: 네이버 공식 쇼핑 API (Client ID/Secret 설정 시)
# 2순위: HTML 파싱 fallback

def _parse_items_from_html(html_text: str, limit: int) -> list[dict]:
    """HTML에서 상품 목록 파싱"""
    from bs4 import BeautifulSoup as _BS
    items = []
    # JSON 상태 파싱
    m = re.search(
        r"window\.__PRELOADED_STATE__\s*=\s*({.+?})\s*(?:</script>|;\s*(?:window|var))",
        html_text, re.S
    )
    if m:
        try:
            state = json.loads(m.group(1))
            prods = (state.get("searchResult",{})
                          .get("shoppingResult",{})
                          .get("products",[]))
            for p in prods[:limit]:
                price = int(re.sub(r"\D","", str(p.get("price","0"))) or "0")
                if price > 0:
                    items.append({
                        "name":    p.get("productName",""),
                        "price":   price,
                        "seller":  p.get("mallName",""),
                        "url":     p.get("mallProductUrl",""),
                        "reviews": str(p.get("reviewCount","")),
                        "rating":  "",
                    })
            if items:
                return items
        except:
            pass
    # HTML fallback
    soup = _BS(html_text, "html.parser")
    for card in soup.select(
        "div.product_item__MDtDF, li.product_item, div[class*=product_item]"
    )[:limit]:
        try:
            nm = card.select_one("a.product_link__TrAac, a.product_link, .name, h2")
            pr = card.select_one("span.price_num__S2p_v, span.price_num, .price")
            ml = card.select_one("span.product_mall__DmHft, .mall_name")
            rv = card.select_one("span.product_grade__H1bUc, .grade_num")
            name  = nm.get_text(strip=True) if nm else ""
            price = int(re.sub(r"\D","", pr.get_text()) or "0") if pr else 0
            if name and price > 0:
                items.append({
                    "name": name, "price": price,
                    "seller": ml.get_text(strip=True) if ml else "",
                    "url": "", "reviews": rv.get_text(strip=True) if rv else "",
                    "rating": "",
                })
        except:
            continue
    return items


def search_naver_shopping(keyword: str, limit: int = 100,
                          p_min: int = 0, p_max: int = 0) -> list[dict]:
    """
    네이버 쇼핑 검색
    p_min/p_max 가격 범위로 필터링해서 limit 개수까지 수집.
    """
    import urllib.parse, time as _time

    cfg = load_config()
    client_id     = cfg.get("naver_client_id", "").strip()
    client_secret = cfg.get("naver_client_secret", "").strip()

    # 방법 1: 공식 쇼핑 API (페이징으로 다건 수집)
    if client_id and client_secret:
        try:
            url     = "https://openapi.naver.com/v1/search/shop.json"
            headers = {
                "X-Naver-Client-Id":     client_id,
                "X-Naver-Client-Secret": client_secret,
                "User-Agent": "Mozilla/5.0",
            }
            collected = []
            start = 1
            max_pages = 10   # 최대 10페이지 (최대 1000개)
            per_page  = 100  # API 한 페이지 최대 100개
            for _ in range(max_pages):
                params = {"query": keyword, "display": per_page,
                          "start": start, "sort": "sim"}
                r = requests.get(url, headers=headers,
                                 params=params, timeout=10)
                if r.status_code != 200:
                    break
                page_items = []
                for p in r.json().get("items", []):
                    price = int(re.sub(r"\D","",
                                str(p.get("lprice","0"))) or "0")
                    name  = re.sub(r"<[^>]+>","", p.get("title",""))
                    page_items.append({
                        "name":    name,
                        "price":   price,
                        "seller":  p.get("mallName",""),
                        "url":     p.get("link",""),
                        "reviews": str(p.get("reviewCount","")),
                        "rating":  str(p.get("grade","")),
                    })
                if not page_items:
                    break

                for item in page_items:
                    price = item["price"]
                    in_range = True
                    if p_min > 0 and price < p_min:
                        in_range = False
                    if p_max > 0 and price > p_max:
                        in_range = False
                    if p_min == 0 and p_max == 0:
                        in_range = True
                    if in_range:
                        collected.append(item)

                if len(collected) >= limit:
                    break
                start += per_page
                _time.sleep(0.3)

            return collected[:limit] if collected else [{"error": "검색 결과 없음"}]
        except Exception:
            pass

    # 방법 2: HTML 파싱 (API 미설정/실패 시)
    headers_crawl = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept":          "text/html,application/xhtml+xml,*/*;q=0.9",
        "Accept-Language": "ko-KR,ko;q=0.9,en;q=0.8",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection":      "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "Sec-Fetch-Dest":  "document",
        "Sec-Fetch-Mode":  "navigate",
        "Sec-Fetch-Site":  "none",
        "Sec-Fetch-User":  "?1",
        "Cache-Control":   "max-age=0",
    }

    try:
        import time as _time2
        sess      = requests.Session()
        base_url  = "https://search.shopping.naver.com/search/all"
        collected = []
        page      = 1
        per_page  = 40
        max_pages = 8   # 최대 8페이지(약 320개) 시도

        while len(collected) < limit and page <= max_pages:
            params = {"query": keyword,
                      "pagingSize": per_page,
                      "pagingIndex": page}
            r = sess.get(base_url, params=params,
                         headers=headers_crawl, timeout=15)

            if r.status_code in (418, 403, 429):
                if page == 1:
                    msg = "네이버 차단 오류 (" + str(r.status_code) + ") - config.json의 naver_client_id / naver_client_secret 설정 필요"
                    return [{"error": msg}]
                break

            if r.status_code != 200:
                break

            page_items = _parse_items_from_html(r.text, per_page)
            if not page_items:
                break

            for item in page_items:
                price = item.get("price", 0)
                ok = True
                if price > 0:
                    if p_min > 0 and price < p_min:
                        ok = False
                    if p_max > 0 and price > p_max:
                        ok = False
                if ok:
                    collected.append(item)

            if len(collected) >= limit:
                break
            page += 1
            _time2.sleep(0.8)

        if collected:
            return collected[:limit]
        return [{"error": "검색 결과를 파싱하지 못했습니다. 네이버 쇼핑 API 설정을 권장합니다."}]

    except Exception as e:
        return [{"error": f"검색 오류: {str(e)[:60]}"}]

def get_my_product_info(url: str) -> dict | None:
    """스마트스토어 상품 정보 파싱"""
    try:
        r = requests.get(url, headers=HEADERS, timeout=10)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")

        # 상품명
        name = ""
        for sel in ["h3.prod_name", "h2.prod_name", "title"]:
            el = soup.select_one(sel)
            if el:
                name = el.get_text(strip=True)
                break

        # 가격
        price = 0
        for sel in ["span.price_num", "em.sell_price", "span.num"]:
            el = soup.select_one(sel)
            if el:
                price = int(re.sub(r"\D", "", el.get_text()) or "0")
                if price > 0:
                    break

        return {"name": name, "price": price, "url": url}
    except Exception as e:
        return {"name": "조회 실패", "price": 0, "url": url, "error": str(e)}

# 엑셀 리포트 생성
def save_excel_report(keyword: str, my_price: int, results: list[dict]) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "가격분석"

    # 헤더 스타일
    hdr_fill  = PatternFill("solid", fgColor="03C75A")
    warn_fill = PatternFill("solid", fgColor="FF4747")
    ok_fill   = PatternFill("solid", fgColor="1a3a1a")
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10

    # 타이틀
    ws.merge_cells("A1:E1")
    ws["A1"] = f"네이버 쇼핑 가격 분석 - {keyword}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = f"내 가격 {my_price:,}원 | 검색 {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, color="888888")
    ws.merge_cells("A2:E2")

    # 컬럼 헤더
    headers = ["상품명", "가격", "판매처", "리뷰수", "내 가격대비"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    # 데이터
    for row_idx, item in enumerate(results, 4):
        price = item.get("price", 0)
        diff  = price - my_price if my_price > 0 and price > 0 else None

        ws.cell(row=row_idx, column=1, value=item.get("name",""))
        ws.cell(row=row_idx, column=2, value=f"{price:,}원" if price else "-")
        ws.cell(row=row_idx, column=3, value=item.get("seller",""))
        ws.cell(row=row_idx, column=4, value=item.get("reviews",""))

        if diff is not None:
            sign = "+" if diff > 0 else ""
            ws.cell(row=row_idx, column=5, value=f"{sign}{diff:,}원")
            if diff < 0:
                for col in range(1, 6):
                    ws.cell(row=row_idx, column=col).fill = warn_fill
                    ws.cell(row=row_idx, column=col).font = Font(color="FFAAAA")
            elif diff > 0:
                for col in range(1, 6):
                    ws.cell(row=row_idx, column=col).fill = ok_fill

    fname = REPORTS_D / f"가격분석_{keyword[:15]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(fname)
    return fname

# 설정 파일 관리
def load_config() -> dict:
    try:
        return json.loads(CONFIG_F.read_text(encoding="utf-8"))
    except:
        return {"products": [], "interval": 60, "alert_sound": True}

def save_config(cfg: dict):
    CONFIG_F.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")

def load_history() -> list:
    try:
        return json.loads(HISTORY_F.read_text(encoding="utf-8"))
    except:
        return []

def append_history(record: dict):
    h = load_history()
    h.insert(0, record)
    HISTORY_F.write_text(json.dumps(h[:200], ensure_ascii=False, indent=2), encoding="utf-8")

# GUI 헬퍼
def mk_label(p, text, font=None, fg=None, bg=None, **kw):
    return tk.Label(p, text=text, font=font or FONT,
                    fg=fg or C["text"], bg=bg or C["panel"], **kw)

def mk_entry(p, var, width=40, show=None):
    e = tk.Entry(p, textvariable=var, width=width,
                 bg=C["input"], fg=C["text"], insertbackground=C["text"],
                 relief="flat", highlightthickness=1,
                 highlightbackground=C["border"], highlightcolor=C["green"],
                 font=FONT, show=show or "")
    return e

def mk_btn(p, text, cmd, color=None, w=None):
    color = color or C["green"]
    b = tk.Button(p, text=text, command=cmd,
                  bg=color, fg="#fff", relief="flat",
                  font=FONT_B, activebackground=C["border"],
                  activeforeground="#fff", cursor="hand2",
                  padx=10, pady=6, **({"width": w} if w else {}))
    b.bind("<Enter>", lambda e: b.config(bg=C["border"]))
    b.bind("<Leave>", lambda e: b.config(bg=color))
    return b

# 라이선스 입력 창
class LicenseWindow(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("라이선스 인증")
        self.geometry("460x290")
        self.resizable(False, False)
        self.configure(bg=C["bg"])
        self.result = None
        self._build()

    def _build(self):
        tk.Frame(self, bg=C["green"], height=3).pack(fill="x")
        wrap = tk.Frame(self, bg=C["bg"])
        wrap.pack(expand=True, fill="both", padx=40, pady=28)

        tk.Label(wrap, text="스마트스토어 가격 모니터",
                 font=FONT_T, bg=C["bg"], fg=C["text"]).pack(pady=(0,4))
        tk.Label(wrap, text="라이선스 키를 입력하세요",
                 font=FONT, bg=C["bg"], fg=C["sub"]).pack(pady=(0,18))

        self.v_key = tk.StringVar()
        e = tk.Entry(wrap, textvariable=self.v_key, width=38,
                     bg=C["input"], fg=C["text"], insertbackground=C["text"],
                     relief="flat", highlightthickness=1,
                     highlightbackground=C["border"], highlightcolor=C["green"],
                     font=("Consolas", 11), justify="center")
        e.pack(fill="x", pady=(0,6))
        e.bind("<Return>", lambda ev: self._verify())

        self.err_lbl = tk.Label(wrap, text="", font=("Malgun Gothic",9),
                                bg=C["bg"], fg=C["red"])
        self.err_lbl.pack(pady=(0,10))

        mk_btn(wrap, "키 인증하기", self._verify, C["green"]).pack(fill="x")
        tk.Label(wrap, text="구매 문의: 크몽 DM",
                 font=("Malgun Gothic",8), bg=C["bg"], fg=C["sub"]).pack(pady=(12,0))

    def _verify(self):
        raw = self.v_key.get().strip().upper()
        # 형식 사전 검증: PLAN-4자리-4자리-8자리-6자리
        import re as _re
        fmt = r'^(BAS|STA|PRE)-[A-Z0-9]{4}-[A-Z0-9]{4}-\d{8}-[A-Z0-9]{6}$'
        if not _re.match(fmt, raw):
            self.err_lbl.config(
                text="키 형식이 올바르지 않습니다. 예: BAS-XXXX-XXXX-20261231-ABCDEF")
            return
        info = verify_key(raw)
        if info is None:
            self.err_lbl.config(text="유효하지 않은 라이선스 키입니다.")
            return
        if not info.get("valid"):
            self.err_lbl.config(text=f"오류: {info.get('reason', '알 수 없음')}")
            return
        if not save_license(raw):
            self.err_lbl.config(text="라이선스 저장 실패. 다시 시도하세요.")
            return
        self.result = info
        self.destroy()

# 메인 앱
class App(tk.Tk):
    def __init__(self, license_info: dict):
        super().__init__()
        self.license   = license_info
        self.max_limit = license_info["limit"]   # 베이직=1, 스탠다드=3

        plan_label = license_info["label"]
        expires    = license_info["expires"]
        self.title(f"스마트스토어 가격 모니터 [{plan_label}]  ~{expires}")
        self.geometry("1000x700")
        self.minsize(900, 600)
        self.configure(bg=C["bg"])

        self.cfg       = load_config()
        self.log_q     = queue.Queue()
        self.running   = False
        self._thread   = None
        self.results   = []

        self._build()
        self._refresh_product_list()
        self._poll()

    # UI 빌드
    def _build(self):
        # 헤더
        hdr = tk.Frame(self, bg="#111111", height=52)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="스마트스토어 가격 모니터",
                 font=FONT_T, bg="#111111", fg=C["text"]).pack(side="left", padx=18)
        self.status_lbl = tk.Label(hdr, text="대기중",
                                   font=FONT_B, bg="#111111", fg=C["sub"])
        self.status_lbl.pack(side="right", padx=18)

        # 라이선스 뱃지
        plan_color = C["blue"] if self.license["plan"] == "STANDARD" else "#aaaaaa"
        tk.Label(hdr, text=f"[ {self.license['label']} | 상품 {self.max_limit}개 | ~{self.license['expires']} ]",
                 font=("Malgun Gothic",9), bg="#111111", fg=plan_color).pack(side="right", padx=(0,12))

        tk.Frame(self, bg=C["green"], height=2).pack(fill="x")

        # 탭 스타일
        style = ttk.Style(self)
        style.theme_use("default")
        style.configure("G.TNotebook", background=C["bg"], borderwidth=0)
        style.configure("G.TNotebook.Tab", background=C["panel"],
                        foreground=C["sub"], padding=[14,7], font=FONT)
        style.map("G.TNotebook.Tab",
                  background=[("selected", C["card"])],
                  foreground=[("selected", C["green"])])

        nb = ttk.Notebook(self, style="G.TNotebook")
        nb.pack(fill="both", expand=True)

        self.t1 = tk.Frame(nb, bg=C["card"])
        self.t2 = tk.Frame(nb, bg=C["card"])
        self.t3 = tk.Frame(nb, bg=C["card"])
        self.t4 = tk.Frame(nb, bg=C["card"])

        nb.add(self.t1, text="상품 관리")
        nb.add(self.t2, text="검색/비교")
        nb.add(self.t3, text="리포트")
        nb.add(self.t4, text="설정")

        self._build_tab1()
        self._build_tab2()
        self._build_tab3()
        self._build_tab4()

        # 하단 바
        bar = tk.Frame(self, bg=C["card"], height=26)
        bar.pack(fill="x", side="bottom")
        bar.pack_propagate(False)
        self.bar = tk.Label(bar, text=f"폴더: {BASE}", font=("Malgun Gothic",8),
                            bg=C["card"], fg=C["sub"])
        self.bar.pack(side="left", padx=8)

    # 탭: 상품 관리
    def _build_tab1(self):
        p = self.t1
        left  = tk.Frame(p, bg=C["card"], width=380)
        right = tk.Frame(p, bg=C["panel"])
        left.pack(side="left", fill="y", padx=0)
        right.pack(side="left", fill="both", expand=True)
        left.pack_propagate(False)

        # 상품 추가 폼
        form = tk.Frame(left, bg=C["card"])
        form.pack(padx=20, pady=20, fill="x")

        tk.Label(form, text="상품 추가", font=FONT_B,
                 bg=C["card"], fg=C["green"]).pack(anchor="w", pady=(0,12))

        self.v_kw    = tk.StringVar()
        self.v_url   = tk.StringVar()
        self.v_price = tk.StringVar()

        for label, var, ph in [
            ("검색 키워드*", self.v_kw,    "예: 무선 블루투스 이어폰"),
            ("내 상품 URL",   self.v_url,   "https://smartstore.naver.com/..."),
            ("내 판매가 (원)", self.v_price, "예: 25000"),
        ]:
            tk.Label(form, text=label, font=("Malgun Gothic",9),
                     bg=C["card"], fg=C["sub"]).pack(anchor="w", pady=(8,3))
            e = mk_entry(form, var)
            e.pack(fill="x")
            e.insert(0, ph)
            e.config(fg=C["sub"])
            def _clear(ev, e=e, ph=ph):
                if e.get() == ph:
                    e.delete(0,"end"); e.config(fg=C["text"])
            def _restore(ev, e=e, ph=ph, var=var):
                if not e.get():
                    e.insert(0, ph); e.config(fg=C["sub"])
            e.bind("<FocusIn>", _clear)
            e.bind("<FocusOut>", _restore)

        tk.Label(form, text="자동 감시 주기 (분)",
                 font=("Malgun Gothic",9), bg=C["card"], fg=C["sub"]).pack(anchor="w", pady=(8,3))
        self.v_interval = tk.StringVar(value=str(self.cfg.get("interval",60)//60))
        mk_entry(form, self.v_interval, width=10).pack(anchor="w")

        mk_btn(form, "상품 추가", self._add_product, C["green"]).pack(
            fill="x", pady=(14,0))

        # 구분선
        tk.Frame(left, bg=C["border"], height=1).pack(fill="x")

        # 감시 제어
        ctrl = tk.Frame(left, bg=C["card"])
        ctrl.pack(padx=20, pady=16, fill="x")
        self.start_btn = mk_btn(ctrl, "감시 시작", self._start_monitor, C["green"])
        self.start_btn.pack(fill="x", pady=(0,6))
        self.stop_btn  = mk_btn(ctrl, "감시 중지", self._stop_monitor, C["red"])
        self.stop_btn.pack(fill="x")
        self.stop_btn.config(state="disabled")

        # 상품 목록 (우측)
        self.list_title = tk.Label(right,
                 text=f"등록된 상품 (0 / {self.max_limit})",
                 font=FONT_B, bg=C["panel"], fg=C["green"])
        self.list_title.pack(anchor="w", padx=16, pady=(16,8))

        list_wrap = tk.Frame(right, bg=C["panel"])
        list_wrap.pack(fill="both", expand=True, padx=16, pady=(0,16))
        sb = tk.Scrollbar(list_wrap, bg=C["card"])
        self.prod_list = tk.Listbox(list_wrap,
                                    bg=C["input"], fg=C["text"],
                                    selectbackground=C["green"], selectforeground="#fff",
                                    font=FONT, relief="flat", borderwidth=0,
                                    activestyle="none",
                                    yscrollcommand=sb.set)
        sb.config(command=self.prod_list.yview)
        sb.pack(side="right", fill="y")
        self.prod_list.pack(fill="both", expand=True)

        mk_btn(right, "선택 상품 삭제", self._del_product, "#333").pack(
            padx=16, pady=(0,8), anchor="e")

    # 탭: 검색/비교
    def _build_tab2(self):
        p = self.t2
        top = tk.Frame(p, bg=C["card"])
        top.pack(fill="x", padx=20, pady=16)

        tk.Label(top, text="키워드 검색", font=FONT_B,
                 bg=C["card"], fg=C["green"]).pack(anchor="w", pady=(0,8))

        row1 = tk.Frame(top, bg=C["card"])
        row1.pack(fill="x")
        self.v_search_kw = tk.StringVar()
        self.v_my_price  = tk.StringVar()
        self.v_price_min = tk.StringVar()
        self.v_price_max = tk.StringVar()

        tk.Label(row1, text="키워드", bg=C["card"], fg=C["sub"], font=FONT).pack(side="left")
        mk_entry(row1, self.v_search_kw, 22).pack(side="left", padx=(6,12))
        tk.Label(row1, text="내 가격", bg=C["card"], fg=C["sub"], font=FONT).pack(side="left")
        mk_entry(row1, self.v_my_price, 9).pack(side="left", padx=(6,16))
        mk_btn(row1, "검색", self._do_search, C["blue"]).pack(side="left", padx=(0,8))
        mk_btn(row1, "엑셀 저장", self._save_excel, "#555").pack(side="left")

        # 가격 범위 필터 (퀵 버튼 포함)
        row2 = tk.Frame(top, bg=C["card"])
        row2.pack(fill="x", pady=(8,0))
        tk.Label(row2, text="가격 범위:", bg=C["card"], fg=C["sub"], font=FONT).pack(side="left")
        mk_entry(row2, self.v_price_min, 10).pack(side="left", padx=(6,4))
        tk.Label(row2, text="원 ~", bg=C["card"], fg=C["sub"], font=FONT).pack(side="left")
        mk_entry(row2, self.v_price_max, 10).pack(side="left", padx=(4,8))
        tk.Label(row2, text="원 (비워두면 전체)", bg=C["card"], fg=C["sub"],
                 font=("Malgun Gothic",9)).pack(side="left")

        # 빠른 범위 버튼
        for label, mn, mx in [("1만원 이하","0","10000"),
                               ("1~3만원","10000","30000"),
                               ("3~5만원","30000","50000"),
                               ("5만원 이상","50000","")]:
            def _set(mn=mn, mx=mx):
                self.v_price_min.set(mn); self.v_price_max.set(mx)
            tk.Button(row2, text=label, command=_set,
                      bg="#2a2a2a", fg=C["sub"], relief="flat",
                      font=("Malgun Gothic",8), padx=6, pady=2,
                      cursor="hand2",
                      activebackground="#333",
                      activeforeground=C["text"]).pack(side="left", padx=(4,0))

        self.search_status = tk.Label(top, text="", bg=C["card"],
                                      fg=C["sub"], font=("Malgun Gothic",9))
        self.search_status.pack(anchor="w", pady=(8,0))

        # 결과 테이블
        cols = ("상품명", "가격", "판매처", "리뷰", "내가격대비")
        style = ttk.Style()
        style.configure("Dark.Treeview",
                        background=C["input"], foreground=C["text"],
                        fieldbackground=C["input"], rowheight=26,
                        font=FONT, borderwidth=0)
        style.configure("Dark.Treeview.Heading",
                        background=C["card"], foreground=C["green"],
                        font=FONT_B, relief="flat")
        style.map("Dark.Treeview",
                  background=[("selected", C["green"])],
                  foreground=[("selected", "#fff")])

        tree_wrap = tk.Frame(p, bg=C["card"])
        tree_wrap.pack(fill="both", expand=True, padx=20, pady=(0,16))

        vsb = ttk.Scrollbar(tree_wrap, orient="vertical")
        self.tree = ttk.Treeview(tree_wrap, columns=cols, show="headings",
                                  style="Dark.Treeview",
                                  yscrollcommand=vsb.set)
        vsb.config(command=self.tree.yview)
        vsb.pack(side="right", fill="y")
        self.tree.pack(fill="both", expand=True)

        widths = [380, 100, 150, 70, 110]
        for col, w in zip(cols, widths):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=w, anchor="center" if col != "상품명" else "w")

        # 태그 색상
        self.tree.tag_configure("cheap",  background="#2a0a0a", foreground="#ff8888")
        self.tree.tag_configure("same",   background="#1a1a2a", foreground="#aaaaff")
        self.tree.tag_configure("pricey", background="#0a1a0a", foreground="#88cc88")

    # 탭: 리포트
    def _build_tab3(self):
        p = self.t3
        top = tk.Frame(p, bg=C["card"])
        top.pack(fill="x", padx=20, pady=16)
        tk.Label(top, text="저장된 리포트", font=FONT_B,
                 bg=C["card"], fg=C["green"]).pack(side="left")
        mk_btn(top, "새로고침", self._refresh_reports, "#333").pack(side="right")
        mk_btn(top, "폴더 열기", self._open_reports_dir, "#333").pack(side="right", padx=(0,8))

        # 리포트 목록
        wrap = tk.Frame(p, bg=C["card"])
        wrap.pack(fill="both", expand=True, padx=20, pady=(0,16))
        sb = tk.Scrollbar(wrap)
        self.report_list = tk.Listbox(wrap,
                                       bg=C["input"], fg=C["text"],
                                       selectbackground=C["green"],
                                       font=FONT, relief="flat",
                                       borderwidth=0, activestyle="none",
                                       yscrollcommand=sb.set)
        sb.config(command=self.report_list.yview)
        sb.pack(side="right", fill="y")
        self.report_list.pack(fill="both", expand=True)
        self.report_list.bind("<Double-Button-1>", self._open_report)

        tk.Label(p, text="더블클릭으로 엑셀 파일 열기",
                 font=("Malgun Gothic",9), bg=C["card"], fg=C["sub"]).pack(pady=(0,12))

        self._refresh_reports()

        # 로그
        tk.Label(p, text="실행 로그", font=FONT_B,
                 bg=C["card"], fg=C["green"]).pack(anchor="w", padx=20, pady=(8,4))
        log_wrap = tk.Frame(p, bg=C["card"])
        log_wrap.pack(fill="x", padx=20, pady=(0,16))
        sb2 = tk.Scrollbar(log_wrap)
        self.log_box = tk.Text(log_wrap, height=8,
                                bg="#080808", fg="#c0c0c0",
                                font=FONT_M, relief="flat",
                                state="disabled", wrap="word",
                                yscrollcommand=sb2.set)
        sb2.config(command=self.log_box.yview)
        sb2.pack(side="right", fill="y")
        self.log_box.pack(fill="x")
        self.log_box.tag_config("ok",   foreground=C["green"])
        self.log_box.tag_config("warn", foreground=C["yellow"])
        self.log_box.tag_config("err",  foreground=C["red"])
        self.log_box.tag_config("info", foreground=C["sub"])

    # 상품 관리 로직
    def _add_product(self):
        kw    = self.v_kw.get().strip()
        url   = self.v_url.get().strip()
        price = self.v_price.get().strip()

        placeholders = ["예: 무선 블루투스 이어폰",
                        "https://smartstore.naver.com/...",
                        "예: 25000"]
        if kw in placeholders: kw = ""
        if url in placeholders: url = ""
        if price in placeholders: price = ""

        if not kw:
            messagebox.showwarning("입력 오류", "키워드를 입력하세요.")
            return

        # 플랜별 상품 개수 제한 체크
        current_count = len(self.cfg.get("products", []))
        if current_count >= self.max_limit:
            if self.license["plan"] == "BASIC":
                messagebox.showwarning(
                    "상품 추가 불가",
                    f"베이직 플랜은 상품 {self.max_limit}개까지만 감시할 수 있습니다.\n\n"
                    f"스탠다드(3개, 139,000원) 또는 프리미엄(5개, 영구, 359,000원)으로\n"
                    f"업그레이드하면 더 많은 상품을 등록할 수 있습니다."
                )
            elif self.license["plan"] == "STANDARD":
                messagebox.showwarning(
                    "상품 추가 불가",
                    f"스탠다드 플랜은 상품 {self.max_limit}개까지만 감시할 수 있습니다.\n\n"
                    f"프리미엄(5개, 영구, 359,000원)으로 업그레이드하면\n"
                    f"더 많은 상품을 등록할 수 있습니다."
                )
            else:
                messagebox.showwarning(
                    "상품 추가 불가",
                    f"최대 {self.max_limit}개까지 등록 가능합니다.\n기존 상품을 삭제 후 추가해주세요."
                )
            return

        try:
            price_int = int(re.sub(r"\D","",price) or "0")
        except:
            price_int = 0

        product = {
            "keyword": kw, "url": url,
            "my_price": price_int,
            "added": datetime.now().strftime("%Y-%m-%d %H:%M"),
        }
        self.cfg.setdefault("products",[]).append(product)
        self.cfg["interval"] = max(30, int(self.v_interval.get() or "1") * 60)
        save_config(self.cfg)
        self._refresh_product_list()
        self._log(f"상품 추가: {kw} (내 가격 {price_int:,}원)", "ok")

    def _del_product(self):
        sel = self.prod_list.curselection()
        if not sel: return
        products = self.cfg.get("products",[])
        if sel[0] < len(products):
            removed = products.pop(sel[0])
            save_config(self.cfg)
            self._refresh_product_list()
            self._log(f"상품 삭제: {removed['keyword']}", "warn")

    def _refresh_product_list(self):
        self.prod_list.delete(0,"end")
        products = self.cfg.get("products",[])
        for p in products:
            price_str = f"  |  내가격 {p['my_price']:,}원" if p.get("my_price") else ""
            self.prod_list.insert("end", f"  - {p['keyword']}{price_str}")
        # 카운트 업데이트
        count = len(products)
        color = C["red"] if count >= self.max_limit else C["green"]
        self.list_title.config(
            text=f"등록된 상품 ({count} / {self.max_limit})",
            fg=color
        )

    # 검색
    def _do_search(self):
        kw = self.v_search_kw.get().strip()
        if not kw:
            messagebox.showwarning("입력 오류", "키워드를 입력하세요.")
            return
        self.search_status.config(text="검색 중...", fg=C["yellow"])
        self.update()

        def _run():
            my_price  = int(re.sub(r"\D","", self.v_my_price.get() or "0") or "0")
            p_min_str = re.sub(r"[^0-9]","", self.v_price_min.get() or "")
            p_max_str = re.sub(r"[^0-9]","", self.v_price_max.get() or "")
            p_min = int(p_min_str) if p_min_str else 0
            p_max = int(p_max_str) if p_max_str else 0

            # 필터 범위를 검색 함수에 직접 전달해 최대 100개 조회
            results = search_naver_shopping(kw, 100, p_min=p_min, p_max=p_max)
            self.results = results
            self.after(0, lambda: self._show_results(
                self.results, my_price, kw, p_min, p_max))

        threading.Thread(target=_run, daemon=True).start()

    def _show_results(self, results: list, my_price: int, kw: str,
                      p_min: int = 0, p_max: int = 0, warn: str = ""):
        for row in self.tree.get_children():
            self.tree.delete(row)

        if results and "error" in results[0]:
            self.search_status.config(
                text="오류: " + results[0]["error"], fg=C["red"])
            return

        valid = [r for r in results if "error" not in r]
        cheaper = sum(1 for r in valid if r.get("price",0) < my_price > 0)

        # 상태 텍스트 구성
        if warn:
            self.search_status.config(text="주의: " + warn, fg=C["yellow"])
        else:
            range_txt = ""
            if p_min > 0 or p_max > 0:
                lo = f"{p_min:,}원" if p_min else "0원"
                hi = f"{p_max:,}원" if p_max else "상한없음"
                range_txt = f"  |  필터: {lo} ~ {hi}"
            price_txt = ""
            if my_price > 0:
                price_txt = f"  |  내가격 {my_price:,}원보다 저렴한 상품: {cheaper}개"
            self.search_status.config(
                text=f"총 {len(valid)}개 결과{range_txt}{price_txt}",
                fg=C["green"] if cheaper == 0 else C["red"]
            )

        in_range_count = 0
        for r in results:
            if "error" in r:
                continue
            price = r.get("price", 0)
            diff  = price - my_price if my_price > 0 and price > 0 else None
            diff_str = (f"+{diff:,}원" if diff and diff > 0 else
                        f"{diff:,}원" if diff else "-")

            # 가격 범위 체크
            in_range = True
            if price > 0:
                if p_min > 0 and price < p_min:
                    in_range = False
                if p_max > 0 and price > p_max:
                    in_range = False

            if in_range and price > 0 and (p_min > 0 or p_max > 0):
                in_range_count += 1

            # 태그 결정: 범위 조건 + 가격 비교 조합
            if not in_range and (p_min > 0 or p_max > 0):
                tag = "out_range"   # 범위 밖
            elif diff is not None and diff < 0:
                tag = "cheap"       # 내 가격보다 저렴
            elif diff is not None and diff > 0:
                tag = "pricey"      # 내 가격보다 비쌈
            else:
                tag = "in_range" if (p_min > 0 or p_max > 0) else "same"

            self.tree.insert("", "end",
                                     values=(r.get("name","")[:50],
                                     f"{price:,}원" if price else "-",
                                     r.get("seller",""),
                                     r.get("reviews",""),
                                     diff_str),
                             tags=(tag,))

        # 범위 태그 색상
        self.tree.tag_configure("in_range",  background="#0a2a0a", foreground="#aaffaa")
        self.tree.tag_configure("out_range", background="#111111", foreground="#444444")

        # 상태바 업데이트 (범위 카운트 포함)
        if (p_min > 0 or p_max > 0) and not warn:
            lo = f"{p_min:,}원" if p_min else "0원"
            hi = f"{p_max:,}원" if p_max else "상한없음"
            range_txt = f"  |  범위({lo}~{hi}) 내 상품: {in_range_count}개"
            price_txt = f"  |  내가격보다 저렴한 상품: {cheaper}개" if my_price > 0 else ""
            self.search_status.config(
                text=f"총 {len(valid)}개 결과{range_txt}{price_txt}",
                fg=C["green"] if in_range_count > 0 else C["yellow"]
            )

        self._log(f"검색 완료: '{kw}' 총 {len(results)}개 (범위내 {in_range_count}개, 저렴 {cheaper}개)", "ok")

    def _save_excel(self):
        kw = self.v_search_kw.get().strip()
        if not self.results:
            messagebox.showinfo("안내", "먼저 검색을 실행하세요.")
            return
        my_price = int(re.sub(r"\D","", self.v_my_price.get() or "0") or "0")
        path = save_excel_report(kw, my_price, self.results)
        self._log(f"엑셀 저장: {path.name}", "ok")
        self._refresh_reports()
        if messagebox.askyesno("저장 완료", f"저장 완료!\n{path}\n\n지금 열어볼까요?"):
            os.startfile(path)

    # 자동 감시
    def _start_monitor(self):
        if not self.cfg.get("products"):
            messagebox.showwarning("상품 없음", "먼저 상품을 추가하세요.")
            return
        self.running = True
        self.start_btn.config(state="disabled")
        self.stop_btn.config(state="normal")
        self.status_lbl.config(text="감시 중", fg=C["green"])
        self._thread = threading.Thread(target=self._monitor_loop, daemon=True)
        self._thread.start()
        self._log("자동 감시 시작", "ok")

    def _stop_monitor(self):
        self.running = False
        self.start_btn.config(state="normal")
        self.stop_btn.config(state="disabled")
        self.status_lbl.config(text="대기중", fg=C["sub"])
        self._log("감시 중지", "warn")

    def _monitor_loop(self):
        while self.running:
            for product in self.cfg.get("products",[]):
                if not self.running: break
                kw       = product.get("keyword","")
                my_price = product.get("my_price", 0)
                self.log_q.put(("info", f"검색 중: {kw}"))

                results = search_naver_shopping(kw, 20)
                if results and "error" not in results[0] and my_price > 0:
                    cheaper = [r for r in results if r.get("price",0) < my_price]
                    if cheaper:
                        cheapest = min(cheaper, key=lambda x: x["price"])
                        diff = cheapest["price"] - my_price
                        self.log_q.put(("warn",
                            f"주의: '{kw}' 저렴한 상품 {len(cheaper)}개 발견! "
                            f"최저: {cheapest['price']:,}원({diff:,}원)"))
                        # 엑셀 자동 저장
                        path = save_excel_report(kw, my_price, results)
                        self.log_q.put(("ok", f"리포트 저장: {path.name}"))
                        append_history({
                            "time": datetime.now().strftime("%Y-%m-%d %H:%M"),
                            "keyword": kw, "my_price": my_price,
                            "cheapest": cheapest["price"],
                            "cheaper_count": len(cheaper),
                        })
                    else:
                        self.log_q.put(("ok", f"'{kw}' 현재 최저가입니다."))

                time.sleep(5)

            interval = self.cfg.get("interval", 60)
            self.log_q.put(("info", f"{interval//60}분 대기..."))
            for _ in range(interval * 10):
                if not self.running: break
                time.sleep(0.1)

    # 리포트
    def _refresh_reports(self):
        self.report_list.delete(0,"end")
        files = sorted(REPORTS_D.glob("*.xlsx"), reverse=True)
        for f in files:
            size = f.stat().st_size // 1024
            self.report_list.insert("end",
                f"  - {f.name}  ({size}KB)")

    def _open_report(self, event=None):
        sel = self.report_list.curselection()
        if not sel: return
        files = sorted(REPORTS_D.glob("*.xlsx"), reverse=True)
        if sel[0] < len(files):
            os.startfile(files[sel[0]])

    def _open_reports_dir(self):
        os.startfile(REPORTS_D)

    # 로그
    def _log(self, msg: str, tag: str = "info"):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_box.config(state="normal")
        self.log_box.insert("end", f"[{ts}] {msg}\n", tag)
        self.log_box.see("end")
        self.log_box.config(state="disabled")

    def _poll(self):
        try:
            while True:
                tag, msg = self.log_q.get_nowait()
                self._log(msg, tag)
                if tag == "warn":
                    self._refresh_reports()
        except queue.Empty:
            pass
        self.after(300, self._poll)

    # 탭: 설정
    def _build_tab4(self):
        p = self.t4
        inner = tk.Frame(p, bg=C["card"])
        inner.pack(padx=32, pady=24, fill="both", expand=True)

        # 네이버 쇼핑 API 설정
        tk.Label(inner, text="네이버 쇼핑 API 설정",
                 font=FONT_B, bg=C["card"], fg=C["green"]).pack(anchor="w", pady=(0,4))
        tk.Frame(inner, bg=C["border"], height=1).pack(fill="x", pady=(0,14))

        # 안내문
        guide = tk.Frame(inner, bg="#1a2a1a",
                         highlightthickness=1,
                         highlightbackground="#2a4a2a")
        guide.pack(fill="x", pady=(0,16))
        tk.Label(guide,
                 text="네이버 쇼핑 API를 설정하면 418 차단 없이 더 안정적으로 검색됩니다. 무료 한도: 일 25,000회 | 발급: developers.naver.com",
                 font=("Malgun Gothic", 9), bg="#1a2a1a", fg="#88cc88",
                 justify="left").pack(padx=12, pady=10, anchor="w")

        self.v_client_id  = tk.StringVar()
        self.v_client_sec = tk.StringVar()

        for lbl, var, ph, is_secret in [
            ("Client ID",     self.v_client_id,  "발급받은 Client ID",     False),
            ("Client Secret", self.v_client_sec, "발급받은 Client Secret", True),
        ]:
            tk.Label(inner, text=lbl, font=("Malgun Gothic",9),
                     bg=C["card"], fg=C["sub"]).pack(anchor="w", pady=(8,3))
            row = tk.Frame(inner, bg=C["card"])
            row.pack(fill="x")
            e = mk_entry(row, var, 50, show="*" if is_secret else None)
            e.pack(side="left", fill="x", expand=True, padx=(0,8))
            if is_secret:
                def toggle(e=e):
                    e.config(show="" if e.cget("show") == "*" else "*")
                mk_btn(row, "보기", toggle, "#333").pack(side="left")

        # 발급 안내 문구
        tk.Label(inner,
                 text="1) developers.naver.com 에서 애플리케이션 등록  2) 검색 API 선택",
                 font=("Malgun Gothic", 9), bg=C["card"], fg=C["sub"]).pack(
                     anchor="w", pady=(6,0))

        tk.Frame(inner, bg=C["border"], height=1).pack(fill="x", pady=(20,0))

        # API 테스트 + 저장 버튼
        btn_row = tk.Frame(inner, bg=C["card"])
        btn_row.pack(fill="x", pady=(12,0))
        mk_btn(btn_row, "연결 테스트", self._test_api, C["blue"]).pack(side="left", padx=(0,8))
        mk_btn(btn_row, "저장", self._save_api_settings, C["green"]).pack(side="left")
        self.api_result = tk.Label(btn_row, text="", font=("Malgun Gothic",9),
                                   bg=C["card"], fg=C["sub"])
        self.api_result.pack(side="left", padx=12)

        # 현재 상태 표시
        tk.Frame(inner, bg=C["border"], height=1).pack(fill="x", pady=(20,0))
        tk.Label(inner, text="현재 API 상태",
                 font=FONT_B, bg=C["card"], fg=C["green"]).pack(anchor="w", pady=(12,6))
        self.api_status_lbl = tk.Label(inner, text="",
                                        font=("Malgun Gothic",10),
                                        bg=C["card"], fg=C["sub"])
        self.api_status_lbl.pack(anchor="w")

        # 저장된 값 로드
        self._load_api_settings()

        tk.Frame(inner, bg=C["card"]).pack(expand=True)

    def _load_api_settings(self):
        cfg = load_config()
        cid = cfg.get("naver_client_id", "")
        sec = cfg.get("naver_client_secret", "")
        self.v_client_id.set(cid)
        self.v_client_sec.set(sec)
        if cid and sec:
            self.api_status_lbl.config(
                text=f"API 설정됨 (ID: {cid[:8]}...)",
                fg=C["green"])
        else:
            self.api_status_lbl.config(
                text="주의: API 미설정 시 검색이 차단될 수 있습니다.",
                fg=C["yellow"])

    def _save_api_settings(self):
        cid = self.v_client_id.get().strip()
        sec = self.v_client_sec.get().strip()
        cfg = load_config()
        cfg["naver_client_id"]     = cid
        cfg["naver_client_secret"] = sec
        save_config(cfg)
        self.cfg = cfg
        self._load_api_settings()
        messagebox.showinfo("저장 완료", "API 키가 저장되었습니다!")

    def _test_api(self):
        cid = self.v_client_id.get().strip()
        sec = self.v_client_sec.get().strip()
        if not cid or not sec:
            self.api_result.config(text="ID와 Secret을 먼저 입력하세요.", fg=C["red"])
            return
        self.api_result.config(text="테스트 중...", fg=C["sub"])
        self.update()
        def _run():
            try:
                r = requests.get(
                    "https://openapi.naver.com/v1/search/shop.json",
                    headers={
                        "X-Naver-Client-Id":     cid,
                        "X-Naver-Client-Secret": sec,
                    },
                    params={"query": "테스트", "display": 1},
                    timeout=8
                )
                if r.status_code == 200:
                    cnt = len(r.json().get("items", []))
                    self.api_result.config(
                        text=f"연결 성공! (결과 {cnt}개)", fg=C["green"])
                elif r.status_code == 401:
                    self.api_result.config(
                        text="인증 실패: ID/Secret을 확인하세요.", fg=C["red"])
                else:
                    self.api_result.config(
                        text=f"오류 코드: {r.status_code}", fg=C["red"])
            except Exception as e:
                self.api_result.config(text=f"오류: {str(e)[:40]}", fg=C["red"])
        import threading as _th
        _th.Thread(target=_run, daemon=True).start()

    def on_close(self):
        self.running = False
        self.destroy()


# 실행
if __name__ == "__main__":
    # 저장된 라이선스 확인
    license_info = load_license()

    if license_info is None or not license_info.get("valid"):
        # 라이선스 입력 창 표시
        win = LicenseWindow()
        win.mainloop()
        license_info = win.result
        if license_info is None:
            sys.exit(0)

    app = App(license_info)
    app.protocol("WM_DELETE_WINDOW", app.on_close)
    app.mainloop()


